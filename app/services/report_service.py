from io import BytesIO
from datetime import datetime, timedelta, UTC

from openpyxl import Workbook
from openpyxl.styles import Font


class ReportService:
    @staticmethod
    def generate_batch_report(batch) -> tuple[BytesIO, int]:
        wb = Workbook()

        ws = wb.active
        ws.title = "Batch"

        ws["A1"] = "Batch Report"
        ws["A1"].font = Font(bold=True)

        ws["A3"] = "Batch ID"
        ws["B3"] = batch.id

        ws["A4"] = "Batch Number"
        ws["B4"] = batch.batch_number

        ws["A5"] = "Date"
        ws["B5"] = str(batch.batch_date)

        ws_products = wb.create_sheet("Products")

        headers = ["ID", "Code", "Aggregated"]
        for i, h in enumerate(headers, 1):
            ws_products.cell(row=1, column=i, value=h).font = Font(bold=True)

        products = batch.products or []

        for row, p in enumerate(products, start=2):
            ws_products.cell(row=row, column=1, value=p.id)
            ws_products.cell(row=row, column=2, value=p.unique_code)
            ws_products.cell(row=row, column=3, value=p.is_aggregated)

        ws_stats = wb.create_sheet("Stats")

        total = len(products)
        aggregated = sum(1 for p in products if p.is_aggregated)

        ws_stats["A1"] = "Total"
        ws_stats["B1"] = total

        ws_stats["A2"] = "Aggregated"
        ws_stats["B2"] = aggregated

        ws_stats["A3"] = "Rate %"
        ws_stats["B3"] = round((aggregated / total * 100) if total else 0, 2)

        ws_stats["A5"] = "Generated"
        ws_stats["B5"] = datetime.now(UTC).isoformat()

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output, len(output.getvalue())

    @staticmethod
    def build_report_result(file_url: str, file_size: int, expires_in_hours: int = 24) -> dict:
        expires_at = datetime.now(UTC) + timedelta(hours=expires_in_hours)
        return {
            "file_url": file_url,
            "file_size": file_size,
            "expires_at": expires_at.isoformat(),
        }